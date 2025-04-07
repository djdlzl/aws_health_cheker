import pandas as pd
import os
from openpyxl import load_workbook

EXCEL_PATH = os.path.join(os.getcwd(), "aws_events.xlsx")

# 이벤트를 계정별 시트에 저장
def append_event_to_excel_by_sheet(excel_write_lock, client_result):
    sheet_name = client_result.get("name", "Unknown")[:31]  # 엑셀 시트 이름 최대 길이 제한

    records = []

    for event_type, events in client_result.get("events", {}).items():
        for event in events:
            row = {}
            row["이벤트 제목"] = event.get("title", "")

            # 상세 정보 삽입
            for key, val in event.get("details", {}).items():
                row[key] = val

            # 영향받는 리소스 정리
            resources = event.get("affected_resources", [])
            resource_str = "\n".join([
                f"{r.get('text')} ({r.get('link')})" if r.get("link") else r.get("text")
                for r in resources
            ]) if resources else ""
            row["영향받는 리소스 목록"] = resource_str

            records.append(row)

    df = pd.DataFrame(records)

    # 🔐 Lock을 사용해 병렬 접근 방지
    with excel_write_lock:
        if not os.path.exists(EXCEL_PATH):
            # 엑셀 파일이 없으면 새로 생성 + 헤더 포함
            with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # 기존 파일 열기 + 시트 존재 여부 체크
            book = load_workbook(EXCEL_PATH)
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                start_row = sheet.max_row
                header = False
            else:
                start_row = 0
                header = True

            with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=header, startrow=start_row)


# 실패한 계정 로그
def log_failed_client(client_name, error_msg):
    log_path = os.path.join(os.getcwd(), "failed_clients.log")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"{client_name} - 실패 이유: {error_msg}\n")

# 전체 Excel 파일 정렬 + 중복 제거
def clean_excel_file():
    if not os.path.exists(EXCEL_PATH):
        return

    book = load_workbook(EXCEL_PATH)
    writer = pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='w')  # overwrite

    for sheet_name in book.sheetnames:
        df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name, engine='openpyxl')

        # 중복 제거 기준: 이벤트 제목 + 시작 시간 + 고객사명
        subset_keys = [k for k in df.columns if k in ["이벤트 제목", "시작 시간", "고객사명"]]
        df = df.drop_duplicates(subset=subset_keys)

        # 정렬: 고객사명 > 이벤트 제목 > 시작 시간
        sort_keys = [k for k in ["고객사명", "이벤트 제목", "시작 시간"] if k in df.columns]
        df = df.sort_values(by=sort_keys)

        df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.close()
    print("엑셀 정리 완료 (중복 제거 및 정렬)")
